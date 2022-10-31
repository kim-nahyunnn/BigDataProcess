#!/usr/bin/python3

import openpyxl

wb = openpyxl.load_workbook("student.xlsx")
ws = wb['Sheet1']

row_id = 1
total=[]
for row in ws:
	if row_id != 1:
		sum = ws.cell(row=row_id, column=3).value * 0.3
		sum += ws.cell(row=row_id, column=4).value * 0.35
		sum += ws.cell(row=row_id, column=5).value * 0.34
		sum += ws.cell(row=row_id, column=6).value
		ws.cell(row=row_id, column=7).value = sum
		total.append(sum)
	row_id += 1
rank = sorted(total)

list = rank.copy() 
grade = ['A+', 'A0', 'B+', 'B0', 'C+', 'C0']
start = 0
aBig = int(len(rank)//0.15)
a = int(len(rank)//0.3)
bBig = int(len(rank)//0.5)
b = int(len(rank)//0.7)
cBig = int(len(rank)//0.85)
list[:aBig] = grade[0]
list[aBig:a] = "A0"
list[a:bBig] = "B+"
list[bBig:b] = "B0"
list[b:cBig] = "C+"
list[cBig:] = "C0"
row_id = 1
index = 0;
for wor in ws:
	if row_id != 1:
		ws.cell(row=row_id, column=8).value = 'C+'
	index += 1
	row_id += 1

wb.save("student.xlsx")
